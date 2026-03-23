import openpyxl
import json
import os

def extract_compilado_clientes(xlsx_path):
    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} not found")
        return None

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'Compilado Clientes' not in wb.sheetnames:
        print(f"Error: 'Compilado Clientes' sheet not found in {xlsx_path}")
        print(f"Available sheets: {wb.sheetnames}")
        return None

    sheet = wb['Compilado Clientes']
    
    # Get headers
    headers = [cell.value for cell in sheet[1]]
    
    data = []
    # Iterate through rows starting from the second one
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row): # Skip empty rows
            continue
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        data.append(row_dict)
    
    return data

if __name__ == "__main__":
    xlsx_path = "ABM Data Mapping.xlsx"
    data = extract_compilado_clientes(xlsx_path)
    if data:
        with open("compilado_clientes.json", "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"Successfully extracted {len(data)} rows to compilado_clientes.json")
